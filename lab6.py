

import requests
import json
from openpyxl import Workbook, load_workbook
import os
from datetime import datetime
import webbrowser


def Exel_file():
    path = os.getcwd()
    a = os.path.exists("mycrypto_data.xlsx")
    if a == False:
        wb = Workbook()
        wb.save("mycrypto_data.xlsx")
        
  
def Check_my_currencies():   
    currencies = ["btcusd", "btceur", "eurusd", "xrpusd", "xrpeur", "xrpbtc", "ltcusd", 
                  "ltceur", "ltcbtc", "ethusd", "etheur", "ethbtc", "bchusd", "bcheur", "bchbtc"]
    cond = True
    while cond == True:
        print('What currency do you have?',"\n","Available: btcusd, btceur, eurusd, xrpusd, xrpeur, xrpbtc, ltcusd, ltceur, ltcbtc, ethusd, etheur, ethbtc, bchusd, bcheur, bchbtc")
        input1 = input()
        
        print("What amount of this currency do you have?")
        input2 = float(input())

        if input1 not in currencies:
            print("I'm sorry, there's no such pair in the system")
            
        else:
            URL = "https://www.bitstamp.net/api/v2/transactions/"+input1+"/"
            query = {'time':'day'}
            r = requests.get(URL,params=query)
            trans = json.loads(r.text)
            transactions = sorted(trans, key=lambda i: i['date'])
            
            profit_per = (float(transactions[0]['price']) / float(transactions[-1]['price'])) * 100 - 100
            profit_mon = float(transactions[0]['price']) * input2 - float(transactions[-1]['price']) * input2
            
            if profit_per >= 0:
                print("Percentage increase:","+", round(float(profit_per),2), "%")
                print("Since yesterday you've earned:", round(profit_mon, 2))
                
            else:
                print("Percentage decrease:",round(float(profit_per),2), "%")
                print("Since yesterday you've lost", round(profit_mon, 2))
                
            date = datetime.now()
            wb = load_workbook("mycrypto_data.xlsx")
            ws = wb.create_sheet(input1)
            ws.cell(row = 1, column = 1).value = "Date of check"
            ws.cell(row = 2, column = 1).value = str(date)
            ws.cell(row = 1, column = 2).value = "Pair of currencies"
            ws.cell(row = 2, column = 2).value = str(input1)
            ws.cell(row = 1, column = 3).value = "Amount of resource"
            ws.cell(row = 2, column = 3).value = str(input2)
            wb.save("mycrypto_data.xlsx")
            
            print("Do you want to check another currency? Input yes or no")
            input3 = input()
            if input3 == "no":
                cond = False
                print("Thank you for using my programme")
                website = "https://imgur.com/gallery/Uyw52SY"
                webbrowser.open_new(website)
   
                    
Exel_file()
Check_my_currencies()        

    
